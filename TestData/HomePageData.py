import openpyxl


class HomePageData:
    test_HomePage_data = [{"name": "Sunderrajan Srinivasan", "mail": "ssrrajan@gmail.com", "password": "ssrrajan123", "gender": "Male"},{"name": "Pavithra Sunderrajan","mail": "pavit88@gmail.com","password": "pavithra123","gender": "Female"}]

    @staticmethod
    def getTestData(test_case_name):
        book = openpyxl.load_workbook("C:\\Sunder\\pyxlDemo1.xlsx")
        sheet = book.get_sheet_by_name("TestData")

        mxrow = sheet.max_row
        mxcol = sheet.max_column

        Dict = {}

        for i in range(1, mxrow + 1):
            Dict = {}
            if sheet.cell(row=i, column=1).value == test_case_name:
                for k in range(2, mxcol + 1):
                    Dict[sheet.cell(row=1, column=k).value] = sheet.cell(row=i, column=k).value

                return [Dict]